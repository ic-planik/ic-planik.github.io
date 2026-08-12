#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
planik_extract.py — extrator do Dashboard Comercial Planik (v2, agosto/2026)

Diferença central para o robô anterior: os blocos são localizados pelos RÓTULOS
que existem na planilha, não por coordenadas fixas. Se uma aba ganhar linhas,
mudar de ordem ou for renomeada dentro de um padrão conhecido, o extrator segue
funcionando — e, quando não seguir, ele FALHA ALTO, dizendo exatamente o que não
encontrou, em vez de publicar número errado em silêncio.

Uso:
    python planik_extract.py "RELATORIO DE VENDAS 2026 VF.xlsx" [saida.json]
"""

import json
import sys
import unicodedata
from collections import defaultdict

import openpyxl

MESES = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
         "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
MES_ABR = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
           "Jul", "Ago", "Set", "Out", "Nov", "Dez"]


# ----------------------------------------------------------------------------
# utilidades
# ----------------------------------------------------------------------------

class ErroDeEstrutura(Exception):
    """A planilha não tem a forma esperada. Falhar aqui é melhor que publicar errado."""


def txt(v):
    return "" if v is None else str(v).strip()


def num(v, padrao=0.0):
    if isinstance(v, bool) or v is None:
        return padrao
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("%", "")
    if s in ("", "-", "—", "N/A", "#DIV/0!", "#N/A", "#VALOR!", "#REF!"):
        return padrao
    s = s.replace(".", "").replace(",", ".") if s.count(",") == 1 and s.count(".") > 1 else s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return padrao


def chave(s):
    """Normaliza texto para comparação: sem acento, sem espaço, maiúsculo."""
    s = unicodedata.normalize("NFKD", txt(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    for ch in (" ", ".", "-", "\u00b7", "\u2014", "\u2013", "_", ":"):
        s = s.replace(ch, "")
    return s.upper()


def achar_linha(ws, alvo, col=1, ini=1, fim=None, exato=False):
    """Primeira linha cuja célula em `col` casa com `alvo`. None se não achar."""
    alvo_k = chave(alvo)
    fim = fim or ws.max_row
    for r in range(ini, fim + 1):
        c = chave(ws.cell(r, col).value)
        if (c == alvo_k) if exato else (alvo_k in c and c != ""):
            return r
    return None


def exigir_linha(ws, alvo, col=1, ini=1, fim=None, exato=False):
    r = achar_linha(ws, alvo, col, ini, fim, exato)
    if r is None:
        raise ErroDeEstrutura(
            f"[{ws.title}] não encontrei a linha com rótulo {alvo!r} na coluna {col}. "
            f"A aba mudou de estrutura — confira antes de publicar."
        )
    return r


def mapear_meses(ws, linha, col_ini, col_fim=None):
    """{índice do mês (0-11): coluna} lendo os nomes de mês numa linha de cabeçalho."""
    col_fim = col_fim or ws.max_column
    fora = {}
    for c in range(col_ini, col_fim + 1):
        k = chave(ws.cell(linha, c).value)
        for i, m in enumerate(MESES):
            if k == chave(m):
                fora[i] = c
    return fora


# ----------------------------------------------------------------------------
# extrator
# ----------------------------------------------------------------------------

class Extrator:
    def __init__(self, caminho):
        self.wb = openpyxl.load_workbook(caminho, data_only=True)
        self.avisos = []
        self._exigir_abas()
        self.meta = self._metadados()
        self.produtos = self._produtos_canonicos()
        self._norm = {chave(p): p for p in self.produtos}

    def aviso(self, msg):
        self.avisos.append(msg)

    def _exigir_abas(self):
        precisa = ["Base", "Metas 2026", "POR GERENTE", "POR PRODUTO YTD",
                   "Por Canal RESUMIDO", "VSOs 2026", "DESCONTO POR PRODUTO",
                   "ORIGEM POR PRODUTO"]
        faltando = [a for a in precisa if a not in self.wb.sheetnames]
        if faltando:
            raise ErroDeEstrutura(
                "Abas ausentes: " + ", ".join(faltando) +
                ".\nAbas presentes: " + ", ".join(self.wb.sheetnames)
            )

    # -- metadados -----------------------------------------------------------

    def _metadados(self):
        """Lê o bloco auto-descritivo de Metas 2026 (Q1:R5) + mês de referência."""
        ws = self.wb["Metas 2026"]
        meta = {}
        for r in range(1, 12):
            rot, val = txt(ws.cell(r, 17).value), ws.cell(r, 18).value
            if not rot:
                continue
            k = chave(rot)
            if k == "ANOBASE":
                meta["anoBase"] = int(num(val, 2026))
            elif k == "FONTEDOREALIZADO":
                meta["fonteRealizado"] = txt(val)
            elif k.startswith("CANALBLOCO"):
                meta.setdefault("canais", []).append(txt(val))

        meta.setdefault("anoBase", 2026)
        meta.setdefault("canais", ["Salão", "Online", "Parcerias"])

        # mês de referência: Metas 2026 S10, conferido contra Por Canal RESUMIDO
        r_ytd = exigir_linha(ws, "Meta YTD", col=16)
        mes_ref = int(num(ws.cell(r_ytd + 1, 19).value, 0))
        pcr = self.wb["Por Canal RESUMIDO"]
        r_ref = achar_linha(pcr, "Mês de referencia", col=1)
        if r_ref:
            outro = int(num(pcr.cell(r_ref + 1, 2).value, 0))
            if outro and mes_ref and outro != mes_ref:
                self.aviso(
                    f"Mês de referência divergente: 'Metas 2026' diz {mes_ref}, "
                    f"'Por Canal RESUMIDO' diz {outro}. Usando {mes_ref}."
                )
        if not 1 <= mes_ref <= 12:
            raise ErroDeEstrutura("Mês de referência inválido em 'Metas 2026' (S10).")
        meta["mesRef"] = mes_ref
        meta["mesRefNome"] = MESES[mes_ref - 1]

        # data de atualização (canto superior da Base)
        base_a1 = txt(self.wb["Base"].cell(1, 1).value)
        meta["atualizado"] = base_a1.split(":")[-1].strip() if ":" in base_a1 else base_a1
        meta["empresa"] = "Planik Empreendimentos Imobiliários"
        meta["periodo"] = f"Janeiro a {MESES[mes_ref - 1]} de {meta['anoBase']}"
        return meta

    def _produtos_canonicos(self):
        """Nomes oficiais de empreendimento = coluna B do bloco de VGV da Metas 2026."""
        ws = self.wb["Metas 2026"]
        r0 = exigir_linha(ws, "META CORTE - VGV", col=1) + 3
        prods = []
        for r in range(r0, r0 + 40):
            nome = txt(ws.cell(r, 2).value)
            if not nome or chave(nome) in ("TOTAL", "ACUMULADO", "REALIZADO", "ATINGIMENTO"):
                break
            prods.append(nome)
        if len(prods) < 5:
            raise ErroDeEstrutura("Não consegui ler a lista de empreendimentos em 'Metas 2026'.")
        return prods

    def canon(self, nome):
        """Concilia as grafias divergentes entre abas (TAJ/Taj, NUV/Nuv, ...)."""
        return self._norm.get(chave(nome), txt(nome))

    # -- Base ----------------------------------------------------------------

    def base(self):
        ws = self.wb["Base"]
        hdr = {txt(ws.cell(2, c).value): c for c in range(1, ws.max_column + 1)
               if txt(ws.cell(2, c).value)}
        # 'Ato' virou carga do site (botao Geral x Ato pago): se sumir, tem que gritar.
        for obrig in ("Produto", "Valor", "Canal", "Gerente", "Ato"):
            if obrig not in hdr:
                raise ErroDeEstrutura(
                    f"[Base] coluna {obrig!r} não existe mais. Cabeçalho atual: {sorted(hdr)}"
                )

        def cel(r, nome):
            c = hdr.get(nome)
            return ws.cell(r, c).value if c else None

        transacoes = []
        for r in range(3, ws.max_row + 1):
            prod = txt(cel(r, "Produto"))
            if not prod:
                continue
            valor = num(cel(r, "Valor"))

            mi = None
            m2 = txt(cel(r, "MÊS2"))
            if m2 and chave(m2) in [chave(m) for m in MESES]:
                mi = [chave(m) for m in MESES].index(chave(m2))
            else:
                d = cel(r, "Mês")
                if hasattr(d, "month"):
                    mi = d.month - 1

            corretor = txt(cel(r, "Corretor")) or None
            dt = cel(r, "Mês")
            data_iso = dt.strftime("%Y-%m-%d") if hasattr(dt, "year") else None
            transacoes.append({
                "unidade":  txt(cel(r, "Unidade")),
                "produto":  self.canon(prod),
                "status":   txt(cel(r, "Status")),
                # 'Ato' = Pago / Pendente. Alimenta o botao Geral x Ato pago do site.
                "ato":      txt(cel(r, "Ato")) or None,
                "mes":      mi,
                "data":     data_iso,
                "corretor": corretor,
                "gerente":  txt(cel(r, "Gerente")) or "—",
                "canal":    txt(cel(r, "Canal")) or "—",
                "parceiro": txt(cel(r, "Parceiro")) or None,
                "valor":    round(valor, 2),
                "desconto": num(cel(r, "Desconto"), None) if cel(r, "Desconto") is not None else None,
                "metragem": num(cel(r, "Metragem"), None) if cel(r, "Metragem") is not None else None,
                "valorM2":  num(cel(r, "Valor do m²"), None) if cel(r, "Valor do m²") is not None else None,
                "fifith":   txt(cel(r, "Fifith?")) or None,
                "origem":   txt(cel(r, "Origem")) or None,
                "plataforma": txt(cel(r, "Plataforma")) or None,
                "campanha": txt(cel(r, "Campanha")) or None,
                "peso":     num(cel(r, "Peso"), 1),
            })

        com_corretor = sum(1 for t in transacoes if t["corretor"])
        return {
            "transacoes": transacoes,
            "qtd": len(transacoes),
            "real": round(sum(t["valor"] for t in transacoes), 2),
            "coberturaCorretor": round(com_corretor / len(transacoes), 4) if transacoes else 0,
        }

    # -- Metas 2026 ----------------------------------------------------------

    def _bloco_metas(self, titulo, tem_share):
        """Lê um bloco de Metas 2026: produtos x 12 meses + Total/Acum/Real/Ating."""
        ws = self.wb["Metas 2026"]
        r_tit = exigir_linha(ws, titulo, col=1)
        r_hdr = None
        for r in range(r_tit, min(r_tit + 6, ws.max_row) + 1):
            if chave(ws.cell(r, 2).value) == "PRODUTO":
                r_hdr = r
                break
        if r_hdr is None:
            raise ErroDeEstrutura(f"[Metas 2026] bloco {titulo!r}: não achei o cabeçalho 'Produto'.")

        cols = mapear_meses(ws, r_hdr, 3, 14)
        if len(cols) != 12:
            raise ErroDeEstrutura(
                f"[Metas 2026] bloco {titulo!r}: esperava 12 meses no cabeçalho, achei {len(cols)}."
            )

        linhas, r = {}, r_hdr + 1
        while r <= ws.max_row:
            nome = txt(ws.cell(r, 2).value)
            rotulo = chave(ws.cell(r, 1).value)
            if rotulo in ("TOTAL", "ACUMULADO", "REALIZADO", "ATINGIMENTO"):
                linhas[rotulo] = r
                if rotulo == "ATINGIMENTO":
                    break
                r += 1
                continue
            if not nome and not rotulo:
                break
            r += 1

        for obrig in ("TOTAL", "ACUMULADO", "REALIZADO"):
            if obrig not in linhas:
                raise ErroDeEstrutura(f"[Metas 2026] bloco {titulo!r}: falta a linha {obrig}.")

        produtos = []
        for r in range(r_hdr + 1, linhas["TOTAL"]):
            nome = txt(ws.cell(r, 2).value)
            if not nome:
                continue
            item = {
                "produto": self.canon(nome),
                "meta": [round(num(ws.cell(r, cols[i]).value), 2) for i in range(12)],
            }
            if tem_share:
                item["share"] = num(ws.cell(r, 1).value)
            else:
                item["ticketMedio"] = num(ws.cell(r, 1).value)
            item["metaAno"] = round(sum(item["meta"]), 2)
            produtos.append(item)

        pega = lambda rot: [round(num(ws.cell(linhas[rot], cols[i]).value), 2) for i in range(12)] \
            if rot in linhas else [0.0] * 12

        return {
            "produtos": produtos,
            "meta":      pega("TOTAL"),
            "metaAcum":  pega("ACUMULADO"),
            "real":      pega("REALIZADO"),
        }

    def metas(self):
        ws = self.wb["Metas 2026"]
        geral = self._bloco_metas("META CORTE - VGV", tem_share=False)

        canais = {}
        for nome in self.meta["canais"]:
            rotulo = None
            for cand in (f"House - {nome}", nome):
                if achar_linha(ws, cand, col=1) is not None:
                    rotulo = cand
                    break
            if rotulo is None:
                self.aviso(f"[Metas 2026] bloco do canal {nome!r} não encontrado — canal ignorado.")
                continue
            canais[nome] = self._bloco_metas(rotulo, tem_share=True)

        # meta em quantidade (unidades)
        qtd = {}
        r_q = achar_linha(ws, "META CORTE - QTD", col=1)
        if r_q is not None:
            r_hdr = r_q + 2
            cols = mapear_meses(ws, r_hdr, 3, 14)
            for r in range(r_hdr + 1, r_hdr + 40):
                nome = txt(ws.cell(r, 2).value)
                if not nome:
                    break
                qtd[self.canon(nome)] = [int(num(ws.cell(r, cols[i]).value)) for i in range(12)]

        # caixa de KPI (Meta YTD / Gap / Atingimento e o mesmo para o ano)
        r_ytd = exigir_linha(ws, "Meta YTD", col=16)
        r_ano = exigir_linha(ws, "Meta ANO", col=16)
        kpi = {
            "metaYTD":    round(num(ws.cell(r_ytd + 1, 16).value), 2),
            "gapYTD":     round(num(ws.cell(r_ytd + 1, 17).value), 2),
            "atingYTD":   round(num(ws.cell(r_ytd + 1, 18).value), 4),
            "metaAno":    round(num(ws.cell(r_ano + 1, 16).value), 2),
            "gapAno":     round(num(ws.cell(r_ano + 1, 17).value), 2),
            "atingAno":   round(num(ws.cell(r_ano + 1, 18).value), 4),
            "mesesRestantes": int(num(ws.cell(r_ano + 1, 19).value)),
        }

        # campanhas ativas (texto de premiação)
        campanhas = []
        r_c = achar_linha(ws, "CAMPANHAS ATIVAS", col=2)
        if r_c:
            for r in range(r_c + 1, ws.max_row + 1):
                nome = txt(ws.cell(r, 2).value)
                if not nome:
                    break
                campanhas.append({
                    "produto":  nome,
                    "corretor": txt(ws.cell(r, 3).value),
                    "gerente":  txt(ws.cell(r, 13).value),
                })

        return {"geral": geral, "canais": canais, "qtd": qtd,
                "kpi": kpi, "campanhas": campanhas}

    # -- POR GERENTE ---------------------------------------------------------

    def _bloco_gerentes(self, ws, r_hdr, col_nome):
        """Lê um bloco 'Gerentes x meses' (Meta VGV / Real VGV / Qtd / Ating)."""
        cols = mapear_meses(ws, r_hdr, col_nome + 1)
        if not cols:
            return []
        linhas = []
        for r in range(r_hdr + 2, ws.max_row + 1):
            nome = txt(ws.cell(r, col_nome).value)
            if not nome or nome == "-":
                if nome == "-":
                    continue
                break
            if chave(nome) == "TOTAL":
                break
            item = {"nome": nome, "meta": [None] * 12, "real": [None] * 12, "qtd": [None] * 12}
            for i, c in cols.items():
                item["meta"][i] = round(num(ws.cell(r, c).value), 2)
                item["real"][i] = round(num(ws.cell(r, c + 1).value), 2)
                item["qtd"][i] = num(ws.cell(r, c + 2).value)
            # YTD = do início do ano até o mês de referência, igual ao que a
            # própria planilha faz nas colunas BU/BV/BW do bloco de Parcerias.
            ate = self.meta["mesRef"]
            item["metaYTD"] = round(sum(v for v in item["meta"][:ate] if v), 2)
            item["realYTD"] = round(sum(v for v in item["real"][:ate] if v), 2)
            item["qtdYTD"] = round(sum(v for v in item["qtd"][:ate] if v), 2)
            item["ating"] = round(item["realYTD"] / item["metaYTD"], 4) if item["metaYTD"] else None
            linhas.append(item)
        return linhas

    def gerentes(self):
        ws = self.wb["POR GERENTE"]
        blocos = {}

        mapa = [
            ("plkGeral",  "Desempenho de Gerentes PLK Online e Salão", 24),
            ("plkSalao",  "Desempenho de Gerentes PLK Salão",          24),
            ("plkOnline", "Desempenho de Gerentes PLK Online (Pós",    24),
            ("parcerias", "Desempenho de Gerentes Parcerias",          24),
            ("lancadoras", "Desempenho de Lançadora",                  24),
        ]
        for id_, titulo, col in mapa:
            r_tit = achar_linha(ws, titulo, col=col)
            if r_tit is None:
                self.aviso(f"[POR GERENTE] bloco {titulo!r} não encontrado.")
                continue
            r_hdr = None
            for r in range(r_tit, min(r_tit + 6, ws.max_row) + 1):
                k = chave(ws.cell(r, col).value)
                if k in ("GERENTES", "LANCADORA"):
                    r_hdr = r
                    break
            if r_hdr is None:
                self.aviso(f"[POR GERENTE] bloco {titulo!r}: cabeçalho não localizado.")
                continue
            blocos[id_] = self._bloco_gerentes(ws, r_hdr, col)

        # parâmetros: mês de entrada de cada gerente
        entrada = {}
        for titulo in ("MÊS DE ENTRADA DOS GERENTES", "LANÇADORAS (mês de entrada)"):
            r_par = achar_linha(ws, titulo, col=24)
            if not r_par:
                continue
            r = r_par + 1
            if chave(ws.cell(r, 24).value) == "GERENTE":
                r += 1
            for rr in range(r, r + 30):
                nome = txt(ws.cell(rr, 24).value)
                if not nome or chave(nome).startswith("AUXILIAR") or chave(nome).startswith("PARAMETROS"):
                    break
                entrada[nome] = {
                    "bloco": txt(ws.cell(rr, 25).value),
                    "mesEntrada": int(num(ws.cell(rr, 26).value, 1)),
                }
        blocos["parametros"] = entrada
        blocos["ativos"] = sorted(entrada.keys())

        # papéis individuais (gerente Online e diretor de Parcerias)
        papeis = []
        for titulo, papel in [("Meta gerente - Online", "gerente_online"),
                              ("Meta diretor - Parcerias", "diretor_parcerias")]:
            r_t = achar_linha(ws, titulo, col=24)
            if not r_t:
                continue
            for r in range(r_t + 1, r_t + 14):
                nome = txt(ws.cell(r, 24).value)
                if chave(nome) == "TOTAL":
                    break
                if not nome or chave(nome) in ("GERENTE", "DIRETOR", "GERENTEONLINE",
                                               "VGVDIRETORPARCERIAS", "VGVPORGERENTE"):
                    continue
                papeis.append({
                    "nome": nome, "papel": papel,
                    "metaYTD": round(num(ws.cell(r, 25).value), 2),
                    "realYTD": round(num(ws.cell(r, 26).value), 2),
                })
        blocos["papeis"] = papeis

        # tabela de conferência da própria planilha (VGV total do ano por gerente)
        total = []
        r_tv = achar_linha(ws, "Total de vendas", col=1)
        if r_tv:
            for r in range(r_tv + 2, r_tv + 30):
                nome = txt(ws.cell(r, 2).value)
                if not nome:
                    break
                total.append({
                    "nome": nome,
                    "canal": txt(ws.cell(r, 1).value),
                    "vgvAno": round(num(ws.cell(r, 3).value), 2),
                })
        blocos["totalAno"] = total
        return blocos

    # -- POR PRODUTO YTD -----------------------------------------------------

    def desempenho(self):
        ws = self.wb["POR PRODUTO YTD"]
        saida = {}
        mapa = [("Geral", "Vendas - Geral"), ("Salão", "Vendas PLK Salão"),
                ("Online", "Vendas PLK Online"), ("Parcerias", "Vendas Parcerias")]
        for nome, titulo in mapa:
            r_tit = achar_linha(ws, titulo, col=1)
            if r_tit is None:
                self.aviso(f"[POR PRODUTO YTD] bloco {titulo!r} não encontrado.")
                continue
            r_hdr = exigir_linha(ws, "Empreendimento", col=1, ini=r_tit, fim=r_tit + 8)
            cols = mapear_meses(ws, r_hdr, 2)
            ytd = {chave(txt(ws.cell(r_hdr + 1, c).value)): c
                   for c in range(2, ws.max_column + 1)}
            c_meta, c_real, c_qtd = (ytd.get("METAYTD"), ytd.get("REALYTD"), ytd.get("QTDYTD"))

            itens = []
            for r in range(r_hdr + 2, r_hdr + 40):
                nm = txt(ws.cell(r, 1).value)
                if not nm or chave(nm) == "TOTAL":
                    break
                itens.append({
                    "produto": self.canon(nm),
                    "meta":  [round(num(ws.cell(r, cols[i]).value), 2) for i in sorted(cols)],
                    "real":  [round(num(ws.cell(r, cols[i] + 1).value), 2) for i in sorted(cols)],
                    "qtd":   [num(ws.cell(r, cols[i] + 2).value) for i in sorted(cols)],
                    "metaYTD": round(num(ws.cell(r, c_meta).value), 2) if c_meta else None,
                    "realYTD": round(num(ws.cell(r, c_real).value), 2) if c_real else None,
                    "qtdYTD":  num(ws.cell(r, c_qtd).value) if c_qtd else None,
                })
            saida[nome] = itens
        return saida

    # -- Conferência ---------------------------------------------------------

    def conferencia(self, kpi, vso_total):
        """
        Os números que a PLANILHA publica nos blocos-resumo dela.

        Não é redundância com o que o site calcula: o site soma a Base linha a
        linha, a planilha soma pelas tabelas dinâmicas e caixas de KPI dela.
        São dois caminhos independentes até o mesmo número — se divergirem,
        alguém mexeu num lado e não no outro, e é exatamente isso que a tela
        de conferência precisa mostrar antes de alguém levar o dado adiante.

        Tudo é localizado por rótulo, nunca por coordenada fixa: se a planilha
        mudar de forma, isto quebra alto em vez de comparar a célula errada.
        """
        ws = self.wb["Por Canal RESUMIDO"]

        # bloco "VENDAS TOTAIS NO ANO POR CANAL": cabeçalho na linha 2, TOTAL na coluna 1
        r_hdr = exigir_linha(ws, "Canal", col=1, exato=True)
        r_tot = exigir_linha(ws, "TOTAL", col=1, ini=r_hdr, fim=r_hdr + 20, exato=True)

        def coluna(rotulo):
            alvo = chave(rotulo)
            for c in range(1, ws.max_column + 1):
                if chave(ws.cell(r_hdr, c).value) == alvo:
                    return c
            raise ErroDeEstrutura(
                f"[Por Canal RESUMIDO] não achei a coluna {rotulo!r} na linha {r_hdr}. "
                f"O bloco de totais mudou de estrutura."
            )

        def val(rotulo, casas=2):
            v = num(ws.cell(r_tot, coluna(rotulo)).value)
            return round(v, casas)

        return {
            "fonte": "blocos-resumo da planilha (Por Canal RESUMIDO, Metas 2026, VSOs 2026)",
            "vgvYTD":   val("VGV REAL YTD"),
            # QTD YTD sai fracionário de propósito: a planilha já conta Fifith
            # como 0,5 para cada equipe. 359 = 353 inteiras + 6 divididas + 16 vagas fora.
            "qtdYTD":   val("QTD YTD", 2),
            "metaYTD":  val("VGV META YTD"),
            "ating":    val("ATINGIMENTO", 6),
            "metaAno":  kpi["metaAno"],
            "vsoAnoYTD": (vso_total or {}).get("vso"),
            "estoqueInicio": (vso_total or {}).get("inicio"),
        }

    # -- Por Canal RESUMIDO --------------------------------------------------

    def sintese_e_canais(self):
        ws = self.wb["Por Canal RESUMIDO"]

        r_s = exigir_linha(ws, "Produto", col=1, ini=25, exato=True)
        sintese = []
        for r in range(r_s + 1, r_s + 30):
            nm = txt(ws.cell(r, 1).value)
            if not nm or nm.startswith("*"):
                break
            sintese.append({
                "produto": self.canon(nm),
                "mediaM2": num(ws.cell(r, 2).value),
                "vgvTotal": round(num(ws.cell(r, 3).value), 2),
                "vgvVendidoVida": round(num(ws.cell(r, 4).value), 2),
                "vgvDisponivel": round(num(ws.cell(r, 5).value), 2),
                "uniTotal": num(ws.cell(r, 6).value),
                "uniVendidasVida": num(ws.cell(r, 7).value),
                "permutas": num(ws.cell(r, 8).value),
                "uniLivres": num(ws.cell(r, 9).value),
                "pctVgvVendido": num(ws.cell(r, 10).value),
                "pctQtdVendido": num(ws.cell(r, 11).value),
            })

        r_m = achar_linha(ws, "CANAIS DE VENDA POR", col=1)
        matriz = []
        if r_m:
            r_hdr = r_m + 1
            canais_cols = {}
            for c in range(2, ws.max_column + 1):
                nm = txt(ws.cell(r_hdr, c).value)
                if nm and chave(nm) != "VGVMETAYTD":
                    canais_cols[nm] = c
            for r in range(r_hdr + 2, r_hdr + 20):
                nm = txt(ws.cell(r, 1).value)
                if not nm or chave(nm) == "TOTAL":
                    break
                linha = {"produto": self.canon(nm), "canais": {}}
                for canal, c in canais_cols.items():
                    linha["canais"][canal] = {
                        "vgv": round(num(ws.cell(r, c).value), 2),
                        "qtd": num(ws.cell(r, c + 1).value),
                        "share": num(ws.cell(r, c + 2).value),
                    }
                matriz.append(linha)
        return {"sintese": sintese, "canalPorProduto": matriz}

    # -- VSOs ----------------------------------------------------------------

    def vso(self):
        ws = self.wb["VSOs 2026"]
        mensal_prod = defaultdict(lambda: [None] * 12)
        mensal_disp = defaultdict(lambda: [None] * 12)
        mensal_vend = defaultdict(lambda: [None] * 12)
        mensal_total = [None] * 12

        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                k = chave(ws.cell(r, c).value)
                if k not in [chave(m) for m in MESES]:
                    continue
                mi = [chave(m) for m in MESES].index(k)
                if chave(ws.cell(r + 1, c).value) != "EMPREENDIMENTO":
                    continue
                for rr in range(r + 2, ws.max_row + 1):
                    nm = txt(ws.cell(rr, c).value)
                    if not nm:
                        break
                    v = num(ws.cell(rr, c + 3).value, None)
                    v = round(v, 4) if v is not None else None
                    # Guardamos também disponível e vendido do mês, não só a razão.
                    # Sem isso não dá para montar a VSO de um subconjunto de produtos
                    # (linha NIK, médio alto, alto padrão): média de VSO não é VSO.
                    disp = num(ws.cell(rr, c + 1).value, None)
                    vend = num(ws.cell(rr, c + 2).value, None)
                    if chave(nm) == "TOTAL":
                        mensal_total[mi] = v
                        break
                    p = self.canon(nm)
                    mensal_prod[p][mi] = v
                    if disp is not None:
                        mensal_disp[p][mi] = round(disp, 2)
                    if vend is not None:
                        mensal_vend[p][mi] = round(vend, 2)

        anual = {}
        r_a = achar_linha(ws, "VSO ANO YTD", col=16)
        if r_a:
            for r in range(r_a + 2, r_a + 20):
                nm = txt(ws.cell(r, 16).value)
                if not nm:
                    break
                anual[self.canon(nm)] = {
                    "inicio": round(num(ws.cell(r, 17).value), 2),
                    "vendas": round(num(ws.cell(r, 18).value), 2),
                    "vso": round(num(ws.cell(r, 19).value), 4),
                }
        return {"mensalPorProduto": dict(mensal_prod),
                "mensalDisponivel": dict(mensal_disp),
                "mensalVendido": dict(mensal_vend),
                "mensalTotal": mensal_total, "anoYTD": anual}

    # -- desconto ------------------------------------------------------------

    def desconto(self):
        ws = self.wb["DESCONTO POR PRODUTO"]
        # exato=True: sem isso o título "DESCONTO POR PRODUTO" casaria com "Produto"
        r_hdr = exigir_linha(ws, "Produto", col=1, exato=True)
        cols = mapear_meses(ws, r_hdr, 2)
        extras = {chave(txt(ws.cell(r_hdr, c).value)): c for c in range(2, ws.max_column + 1)}
        itens = []
        for r in range(r_hdr + 1, r_hdr + 30):
            nm = txt(ws.cell(r, 1).value)
            if not nm:
                break
            pega = lambda k: round(num(ws.cell(r, extras[k]).value, None), 4) \
                if k in extras and ws.cell(r, extras[k]).value is not None else None
            itens.append({
                "produto": self.canon(nm),
                "mensal": [round(num(ws.cell(r, cols[i]).value, None), 4)
                           if ws.cell(r, cols[i]).value not in (None, "-") else None
                           for i in sorted(cols)],
                "media": pega("MEDIATOTAL"),
                "maximo": pega("DESCONTOMAXIMO"),
                "perda": pega("PERDA"),
                "perdaNominal": pega("PERDANOMINAL"),
            })
        return itens

    # -- origem (marketing) --------------------------------------------------

    def origem(self):
        ws = self.wb["ORIGEM POR PRODUTO"]
        r_hdr = exigir_linha(ws, "Empreendimento", col=1)
        detalhe = []
        for r in range(r_hdr + 1, ws.max_row + 1):
            nm = txt(ws.cell(r, 1).value)
            if not nm:
                break
            detalhe.append({
                "produto":  self.canon(nm),
                "mes":      txt(ws.cell(r, 2).value),
                "canal":    txt(ws.cell(r, 3).value),
                "origem":   txt(ws.cell(r, 4).value) or "Não informada",
                "plataforma": txt(ws.cell(r, 5).value) or None,
                "campanha": txt(ws.cell(r, 6).value) or None,
                "vendas":   num(ws.cell(r, 7).value),
            })

        ranking = []
        r_rk = achar_linha(ws, "Ranking de Origens", col=13)
        if r_rk:
            for r in range(r_rk + 2, ws.max_row + 1):
                nm = txt(ws.cell(r, 13).value)
                if not nm:
                    break
                ranking.append({"origem": nm, "vendas": num(ws.cell(r, 14).value)})

        por_canal = []
        r_pc = achar_linha(ws, "Por canal", col=9)
        if r_pc:
            for r in range(r_pc + 2, r_pc + 15):
                nm = txt(ws.cell(r, 9).value)
                if not nm or chave(nm) == "TOTAL":
                    break
                por_canal.append({
                    "canal": nm,
                    "vendas": num(ws.cell(r, 10).value),
                    "share": num(ws.cell(r, 11).value),
                })
        return {"detalhe": detalhe, "rankingOrigens": ranking, "porCanal": por_canal}

    def origem_da_base(self, transacoes):
        """
        Agregados de marketing calculados a partir da Base.

        Por que não usar a aba ORIGEM POR PRODUTO: ela conta vendas mas não traz
        VGV, e deixa de fora as vendas sem origem preenchida e as de origem
        'Vaga' — 16 vendas ponderadas que somem do ranking. A Base tem as mesmas
        colunas (Origem, Plataforma, Campanha) em todas as linhas, então dá para
        cruzar origem com VGV e abrir a venda individual no clique.
        """
        def agrupa(chave_fn):
            acc = defaultdict(lambda: {"qtd": 0.0, "linhas": 0, "vgv": 0.0, "canais": defaultdict(float)})
            for t in transacoes:
                k = chave_fn(t)
                if k is None:
                    continue
                a = acc[k]
                a["qtd"] += t["peso"]
                a["linhas"] += 1
                a["vgv"] += t["valor"]
                a["canais"][t["canal"]] += t["valor"]
            saida = []
            for k, a in acc.items():
                saida.append({
                    "nome": k,
                    "qtd": round(a["qtd"], 2),
                    "vendas": a["linhas"],
                    "vgv": round(a["vgv"], 2),
                    "ticket": round(a["vgv"] / a["linhas"], 2) if a["linhas"] else 0,
                    "canais": {c: round(v, 2) for c, v in a["canais"].items()},
                })
            return sorted(saida, key=lambda x: -x["vgv"])

        total_vgv = sum(t["valor"] for t in transacoes) or 1
        por_canal = agrupa(lambda t: t["canal"])
        for c in por_canal:
            c["share"] = round(c["vgv"] / total_vgv, 4)
        ranking = agrupa(lambda t: t["origem"] or "Não informada")
        for o in ranking:
            o["share"] = round(o["vgv"] / total_vgv, 4)

        sem_origem = sum(t["peso"] for t in transacoes
                         if not t["origem"] or chave(t["origem"]) in ("VAGA", "NAOINFORMADA"))
        return {
            "porCanal": por_canal,
            "rankingOrigens": ranking,
            "plataformas": agrupa(lambda t: t["plataforma"]),
            "campanhas": agrupa(lambda t: t["campanha"]),
            "porProduto": agrupa(lambda t: t["produto"]),
            "qualidade": {
                "semOrigem": round(sem_origem, 2),
                "totalPonderado": round(sum(t["peso"] for t in transacoes), 2),
                "pctSemOrigem": round(sem_origem / sum(t["peso"] for t in transacoes), 4) if transacoes else 0,
            },
        }

    def m2_por_produto(self, transacoes):
        """
        Valor do m² médio de assinatura: VGV assinado ÷ metragem assinada, por produto.

        Dois cuidados que mudam bastante o resultado:

        1. Venda dividida (coluna 'Fifith?' = Sim). A planilha grava DUAS linhas para
           a mesma unidade, uma por gerente, com metade do valor em cada — mas repete
           a metragem cheia nas duas. Somar as duas linhas direto dobraria a metragem
           e derrubaria o valor do m² pela metade. Aqui as linhas da mesma unidade são
           unificadas: soma-se o valor e conta-se a metragem uma vez só.

        2. Vagas de garagem (peso 0). Têm metragem e preço por m² muito diferentes de
           uma unidade e não contam como unidade vendida — ficam de fora da média.
        """
        grupos = defaultdict(lambda: {"vgv": 0.0, "m2": 0.0, "linhas": 0})
        for t in transacoes:
            if not t.get("metragem") or not t.get("peso"):
                continue
            g = grupos[(t["produto"], t.get("unidade") or f"_{id(t)}")]
            g["vgv"] += t["valor"]
            g["m2"] = max(g["m2"], t["metragem"])   # a metragem se repete nas linhas divididas
            g["linhas"] += 1

        acc = defaultdict(lambda: {"vgv": 0.0, "m2": 0.0, "n": 0})
        for (prod, _uni), g in grupos.items():
            a = acc[prod]
            a["vgv"] += g["vgv"]; a["m2"] += g["m2"]; a["n"] += 1

        saida = {}
        for p, a in acc.items():
            if a["m2"] > 0:
                saida[p] = {"valorM2": round(a["vgv"] / a["m2"], 2),
                            "metragemMedia": round(a["m2"] / a["n"], 2),
                            "vendas": a["n"]}
        tot_v = sum(a["vgv"] for a in acc.values())
        tot_m = sum(a["m2"] for a in acc.values())
        saida["_geral"] = {"valorM2": round(tot_v / tot_m, 2) if tot_m else None,
                           "metragemMedia": None, "vendas": sum(a["n"] for a in acc.values())}
        return saida

    # -- montagem ------------------------------------------------------------

    def montar(self):
        base = self.base()
        metas = self.metas()
        payload = {
            "_schema": "planik.dashboard/2",
            "meta": self.meta,
            "kpi": metas["kpi"],
            "realYTD": base["real"],
            "qtdYTD": base["qtd"],
            "coberturaCorretor": base["coberturaCorretor"],
            "metas": metas,
            "gerentes": self.gerentes(),
            "desemp": self.desempenho(),
            "vso": self.vso(),
            "desconto": self.desconto(),
            "origem": self.origem(),
            "mkt": self.origem_da_base(base["transacoes"]),
            "m2": self.m2_por_produto(base["transacoes"]),
            "transacoes": base["transacoes"],
        }
        payload.update(self.sintese_e_canais())
        payload["conferencia"] = self.conferencia(
            metas["kpi"], (payload["vso"].get("anoYTD") or {}).get("TOTAL"))
        payload["_avisos"] = self.avisos
        return payload


# ----------------------------------------------------------------------------
# verificação
# ----------------------------------------------------------------------------

def verificar(p):
    """Confere o extraído contra as próprias células de conferência da planilha."""
    linhas, erros = [], []
    m = p["meta"]
    k = p["kpi"]
    mesref = m["mesRef"]

    linhas.append(f"Planilha atualizada em {m['atualizado']} · mês de referência: {m['mesRefNome']} ({mesref})")
    linhas.append(f"Realizado YTD (Base) ..... R$ {p['realYTD']:,.2f} em {p['qtdYTD']} vendas")
    linhas.append(f"Meta YTD ................. R$ {k['metaYTD']:,.2f}")
    linhas.append(f"Atingimento YTD .......... {k['atingYTD']*100:,.2f}%")
    linhas.append(f"Meta ANO ................. R$ {k['metaAno']:,.2f}  (atingimento {k['atingAno']*100:,.2f}%)")

    real_metas = sum(p["metas"]["geral"]["real"])
    if abs(real_metas - p["realYTD"]) > 1:
        erros.append(f"Realizado da Base ({p['realYTD']:,.2f}) diverge do da Metas 2026 ({real_metas:,.2f}).")

    ating = p["realYTD"] / k["metaYTD"] if k["metaYTD"] else 0
    if abs(ating - k["atingYTD"]) > 0.001:
        erros.append(f"Atingimento recalculado ({ating:.4f}) diverge do da planilha ({k['atingYTD']:.4f}).")

    linhas.append("")
    linhas.append("Canais (realizado YTD):")
    soma = 0
    for nome, bloco in p["metas"]["canais"].items():
        rl, mt = sum(bloco["real"]), bloco["metaAcum"][mesref - 1]
        soma += rl
        linhas.append(f"  {nome:<10} real R$ {rl:>16,.2f} | meta YTD R$ {mt:>16,.2f} | {rl/mt*100 if mt else 0:6.2f}%")
    linhas.append(f"  {'Interna':<10} real R$ {p['realYTD']-soma:>16,.2f} | (sem meta)")

    linhas.append("")
    linhas.append("Gerentes — Parcerias:")
    for g in p["gerentes"].get("parcerias", []):
        at = f"{g['ating']*100:6.1f}%" if g["ating"] else "   —  "
        linhas.append(f"  {g['nome']:<10} real R$ {g['realYTD']:>15,.2f} | meta R$ {g['metaYTD']:>15,.2f} | {at}")
    linhas.append("Gerentes — PLK (Salão + Online):")
    for g in p["gerentes"].get("plkGeral", []):
        at = f"{g['ating']*100:6.1f}%" if g["ating"] else "   —  "
        linhas.append(f"  {g['nome']:<10} real R$ {g['realYTD']:>15,.2f} | meta R$ {g['metaYTD']:>15,.2f} | {at}")

    # confronto: VGV do ano por gerente (planilha) x soma da Base
    porger = defaultdict(float)
    for t in p["transacoes"]:
        porger[t["gerente"]] += t["valor"]
    if p["gerentes"].get("totalAno"):
        linhas.append("")
        linhas.append("VGV do ano por gerente — planilha x Base:")
        for g in p["gerentes"]["totalAno"]:
            b = porger.get(g["nome"], 0.0)
            marca = "ok" if abs(b - g["vgvAno"]) < 1 else "DIVERGE"
            linhas.append(f"  {g['nome']:<10} planilha R$ {g['vgvAno']:>15,.2f} | Base R$ {b:>15,.2f}  {marca}")
            if marca == "DIVERGE":
                erros.append(f"VGV de {g['nome']}: planilha {g['vgvAno']:,.2f} x Base {b:,.2f}")

    linhas.append("")
    linhas.append(f"Empreendimentos ......... {len(p['metas']['geral']['produtos'])}")
    linhas.append(f"Transações .............. {len(p['transacoes'])} (cobertura de corretor {p['coberturaCorretor']*100:.0f}%)")
    linhas.append(f"Linhas de origem (MKT) .. {len(p['origem']['detalhe'])}")
    linhas.append(f"Ranking de origens ...... {len(p['origem']['rankingOrigens'])}")
    linhas.append(f"Campanhas ativas ........ {len(p['metas']['campanhas'])}")
    linhas.append(f"Papéis individuais ...... {[x['nome'] for x in p['gerentes'].get('papeis', [])]}")

    return "\n".join(linhas), erros


def main():
    entrada = sys.argv[1] if len(sys.argv) > 1 else "planilha.xlsx"
    saida = sys.argv[2] if len(sys.argv) > 2 else "payload.json"
    try:
        p = Extrator(entrada).montar()
    except ErroDeEstrutura as e:
        print("FALHA DE ESTRUTURA — nada foi publicado.\n", e, file=sys.stderr)
        sys.exit(2)

    relatorio, erros = verificar(p)
    print(relatorio)
    if p["_avisos"]:
        print("\nAvisos:")
        for a in p["_avisos"]:
            print("  •", a)
    if erros:
        print("\nDIVERGÊNCIAS:", file=sys.stderr)
        for e in erros:
            print("  ✗", e, file=sys.stderr)

    with open(saida, "w", encoding="utf-8") as f:
        json.dump(p, f, ensure_ascii=False, separators=(",", ":"))
    print(f"\nPayload gravado em {saida} ({len(json.dumps(p, ensure_ascii=False)):,} caracteres)")
    sys.exit(1 if erros else 0)


if __name__ == "__main__":
    main()
